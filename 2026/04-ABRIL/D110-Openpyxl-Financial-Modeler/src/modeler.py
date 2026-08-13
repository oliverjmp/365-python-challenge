import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

class FinancialModeler:
    def __init__(self, data: pd.DataFrame):
        self.data = data

    def generate_excel_model(self, filepath: str):
        """Crea un modelo financiero en Excel estructurado con fórmulas y formatos condicionales."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modelo Financiero"

        # Asegurar que las cuadrículas estén visibles
        ws.views.sheetView[0].showGridLines = True

        # Estilos corporativos
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        bold_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        total_top_border = Side(style='thin', color='000000')
        total_bottom_border = Side(style='double', color='000000')
        total_border = Border(top=total_top_border, bottom=total_bottom_border)

        # Título del reporte
        ws["A1"] = "Reporte de Modelo Financiero Proyectado"
        ws["A1"].font = title_font
        ws.append([]) # Fila en blanco

        # Cabeceras de la tabla
        headers = ["Concepto", "Mes 1", "Mes 2", "Mes 3", "Total"]
        ws.append(headers)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Inserción de datos y fórmulas por fila
        start_row = 4
        for i, row_data in self.data.iterrows():
            concept = row_data["Concepto"]
            m1 = row_data["Mes 1"]
            m2 = row_data["Mes 2"]
            m3 = row_data["Mes 3"]
            
            current_row = start_row + i
            # Fórmula de Excel para sumar el total de la fila (ej. =SUM(B4:D4))
            total_formula = f"=SUM(B{current_row}:D{current_row})"
            
            row_values = [concept, m1, m2, m3, total_formula]
            ws.append(row_values)

            # Aplicar bordes y formatos numéricos
            for c in range(1, 6):
                cell = ws.cell(row=current_row, column=c)
                cell.border = thin_border
                if c > 1:
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # Fila de Totales Generales con Fórmulas
        last_data_row = start_row + len(self.data) - 1
        total_row_idx = last_data_row + 1
        
        sum_col_b = f"=SUM(B{start_row}:B{last_data_row})"
        sum_col_c = f"=SUM(C{start_row}:C{last_data_row})"
        sum_col_d = f"=SUM(D{start_row}:D{last_data_row})"
        sum_col_e = f"=SUM(E{start_row}:E{last_data_row})"

        ws.cell(row=total_row_idx, column=1, value="TOTAL GENERAL").font = bold_font
        ws.cell(row=total_row_idx, column=2, value=sum_col_b).font = bold_font
        ws.cell(row=total_row_idx, column=3, value=sum_col_c).font = bold_font
        ws.cell(row=total_row_idx, column=4, value=sum_col_d).font = bold_font
        ws.cell(row=total_row_idx, column=5, value=sum_col_e).font = bold_font

        for c in range(1, 6):
            cell = ws.cell(row=total_row_idx, column=c)
            cell.border = total_border
            if c > 1:
                cell.number_format = '$#,##0.00'

        # Formato Condicional: Resaltar valores mayores a 5000 en verde claro
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        green_font = Font(color='006100')
        ws.conditional_formatting.add(
            f'B4:E{last_data_row}',
            CellIsRule(operator='greaterThan', formula=['5000'], stopIfTrue=True, fill=green_fill, font=green_font)
        )

        # Autoajuste de ancho de columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(filepath)
        return filepath