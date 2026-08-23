import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_styled_excel(df: pd.DataFrame, sheet_name: str = "Reporte") -> io.BytesIO:
    """
    Exporta un DataFrame a Excel aplicando formato corporativo profesional usando openpyxl.
    Retorna un buffer de bytes listo para descargas.
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
    # Recargamos con openpyxl para aplicar estilos avanzados
    output.seek(0)
    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb[sheet_name]
    
    # Definición de estilos profesionales
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Aplicar estilos a la cabecera
    for col_num in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
    # Aplicar estilos y bordes a los datos y ajustar ancho de columnas
    for col_num in range(1, len(df.columns) + 1):
        max_len = len(str(ws.cell(row=1, column=col_num).value))
        col_letter = get_column_letter(col_num)
        
        for row_num in range(2, len(df) + 2):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left" if isinstance(cell.value, str) else "right")
            
            # Evaluar longitud máxima para ajustar ancho
            cell_len = len(str(cell.value or ""))
            if cell_len > max_len:
                max_len = cell_len
                
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output