import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import sys
import os

# CONFIGURACIÓN DE RUTAS ABSOLUTAS TOTALES
# Esto detecta la carpeta '2026' o la raíz del proyecto para no perderse
BASE_DIR = Path(__file__).resolve().parent # Carpeta D03
ROOT_DIR = BASE_DIR.parent # Carpeta 01-ENERO 2026 (o similar)

# Buscamos el archivo del día 2 de forma explícita
RUTA_CSV_DIA2 = ROOT_DIR / "D02-Scraper-Avanzado" / "precios_crypto.csv"
ARCHIVO_EXCEL = BASE_DIR / "informe_cripto.xlsx"

def ejecutar_excel():
    try:
        # Verificación de existencia con impresión de ruta para el log
        if not RUTA_CSV_DIA2.exists():
            print(f"ERROR: No se encontro el CSV en la ruta: {RUTA_CSV_DIA2}")
            return False
        
        # 1. Cargar datos
        df = pd.read_csv(RUTA_CSV_DIA2)
        
        # 2. Guardar Excel Base
        # Forzamos el motor 'xlsxwriter' si 'openpyxl' da problemas al escribir
        df.to_excel(ARCHIVO_EXCEL, index=False)
        
        # 3. Abrir con Openpyxl para el diseño
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        
        # Aplicar un formato mínimo para asegurar que no falle por estilos
        azul_relleno = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = azul_relleno
            
        wb.save(ARCHIVO_EXCEL)
        wb.close()
        return True

    except Exception as e:
        print(f"ERROR DETALLADO D03: {str(e)}")
        return False

if __name__ == "__main__":
    # Forzamos que el script sepa dónde está parado
    os.chdir(str(BASE_DIR))
    
    if ejecutar_excel():
        print("D03_SUCCESS")
        sys.exit(0)
    else:
        sys.exit(1)