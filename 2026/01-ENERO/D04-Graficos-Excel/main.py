import pandas as pd
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from pathlib import Path

# ============================================================
# CONFIGURACIÓN DE RUTAS DINÁMICAS
# ============================================================
BASE_DIR = Path(__file__).resolve().parent
# Referencia al archivo generado en el Día 03
RUTA_EXCEL_DIA3 = BASE_DIR.parent / "D03-Excel-Profesional" / "informe_cripto.xlsx"
OUTPUT_EXCEL = BASE_DIR / "informe_cripto_premium.xlsx"

# ============================================================
# ESTILOS PREMIUM
# ============================================================
GRIS_OSCURO = "1C1C1C"
DORADO = "D4AF37"
BLANCO = "FFFFFF"

font_blanca = Font(color=BLANCO, name="Calibri Light", size=11)
font_negra_bold = Font(color="000000", name="Calibri Light", size=12, bold=True)
borde_fino = Border(
    left=Side(style="thin", color="666666"), 
    right=Side(style="thin", color="666666"),
    top=Side(style="thin", color="666666"), 
    bottom=Side(style="thin", color="666666")
)

# ============================================================
# 1. CARGAR Y LIMPIAR DATOS
# ============================================================
def cargar_y_limpiar():
    if not RUTA_EXCEL_DIA3.exists():
        print(f"STDERR: Archivo no encontrado en {RUTA_EXCEL_DIA3}")
        return None
    
    # Leemos intentando mantener tipos numéricos
    df = pd.read_excel(RUTA_EXCEL_DIA3)
    
    # Normalización de nombres (Sincronizado con D03)
    # Si las columnas vienen con guion bajo del D03, nos aseguramos de mapearlas
    df.columns = [c.replace(" ", "_") for c in df.columns]
    
    return df

# ============================================================
# 2. CONSTRUCCIÓN DE HOJAS
# ============================================================

def crear_portada(wb):
    ws = wb.active
    ws.title = "Portada"
    ws.sheet_view.showGridLines = False

    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=10):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="000000")

    fecha = datetime.now().strftime("%d/%m/%Y")
    ws.merge_cells("B5:I7")
    ws["B5"].value = "MERCADO CRIPTO - ANALISIS EJECUTIVO"
    ws["B5"].font = Font(color=BLANCO, name="Calibri Light", size=28, bold=True)
    ws["B5"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B20:I22")
    ws["B20"].value = f"Oliver Javier Morales Perez - Actualizado: {fecha}"
    ws["B20"].font = font_blanca
    ws["B20"].alignment = Alignment(horizontal="center", vertical="center")

def crear_tabla(wb, df):
    ws = wb.create_sheet("Tabla")
    ws.sheet_view.showGridLines = False
    
    columnas = list(df.columns)
    ws.append(columnas)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=DORADO)
        cell.font = font_negra_bold
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde_fino

    for fila in df.values.tolist():
        ws.append(fila)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=GRIS_OSCURO)
            cell.font = font_blanca
            cell.border = borde_fino
            cell.alignment = Alignment(horizontal="center")

    # Formatos (Usando nombres con guion bajo)
    try:
        idx_precio = columnas.index("Precio") + 1
        idx_cambio = columnas.index("Cambio_24h") + 1
        idx_mc = columnas.index("Market_Cap") + 1

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=idx_precio).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=idx_cambio).number_format = '0.00%'
            ws.cell(row=row, column=idx_mc).number_format = '"$"#,##0'
    except Exception as e:
        print(f"Aviso Formato: {e}")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

def crear_marketcap_sheet(wb, df):
    ws = wb.create_sheet("MarketCap")
    ws.sheet_view.showGridLines = False
    ws.append(["Criptomoneda", "Market_Cap"])
    
    for _, row in df.iterrows():
        ws.append([row["Nombre"], row["Market_Cap"]])

    chart = BarChart()
    chart.title = "Market Cap (USD)"
    chart.width = 30
    chart.height = 15
    
    datos = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(datos, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")

def ejecutar():
    try:
        df_limpio = cargar_y_limpiar()
        if df_limpio is None: return False

        wb = Workbook()
        crear_portada(wb)
        crear_tabla(wb, df_limpio)
        crear_marketcap_sheet(wb, df_limpio)
        
        wb.save(OUTPUT_EXCEL)
        print("D04_OK: Dashboard Premium generado.")
        return True
    except Exception as e:
        print(f"STDERR: Error en D04: {e}")
        return False

if __name__ == "__main__":
    print("Iniciando D04 - Dashboard...")
    if ejecutar():
        sys.exit(0)
    else:
        sys.exit(1)