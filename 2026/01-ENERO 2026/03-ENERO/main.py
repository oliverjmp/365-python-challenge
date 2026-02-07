import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# 1. CARGAR CSV DESDE EL DÍA 02 (RUTA AUTOMÁTICA)
# ============================================================

def cargar_csv():
    # Carpeta donde está este script del Día 03
    carpeta_actual = os.path.dirname(os.path.abspath(__file__))

    # Ruta hacia el CSV del Día 02
    ruta_csv = os.path.join(carpeta_actual, "..", "02-ENERO", "precios_crypto.csv")

    # Normalizar ruta para Windows/Mac/Linux
    ruta_csv = os.path.normpath(ruta_csv)

    print(f"Buscando CSV en: {ruta_csv}")

    df = pd.read_csv(ruta_csv)
    return df

# ============================================================
# 2. EXPORTAR A EXCEL (BASE)
# ============================================================

def exportar_a_excel(df, nombre_archivo="informe_cripto.xlsx"):
    df.to_excel(nombre_archivo, index=False)
    return nombre_archivo

# ============================================================
# 3. APLICAR FORMATO PROFESIONAL CON OPENPYXL
# ============================================================

def formatear_excel(nombre_archivo="informe_cripto.xlsx"):
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    # Encabezados: fondo azul, texto blanco, negrita
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Detectar columnas por nombre
    col_precio = None
    col_cambio = None
    col_marketcap = None

    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Precio":
            col_precio = idx
        elif cell.value == "Cambio 24h":
            col_cambio = idx
        elif cell.value == "Market Cap":
            col_marketcap = idx

    # Formato numérico
    for row in range(2, ws.max_row + 1):

        # Precio → moneda
        if col_precio:
            cell = ws.cell(row=row, column=col_precio)
            try:
                valor = float(str(cell.value).replace("$", "").replace(",", ""))
                cell.value = valor
                cell.number_format = '"$"#,##0.00'
            except:
                pass

        # Cambio 24h → porcentaje
        if col_cambio:
            cell = ws.cell(row=row, column=col_cambio)
            try:
                valor = float(str(cell.value).replace("%", ""))
                cell.value = valor / 100
                cell.number_format = "0.00%"
            except:
                pass

        # Market Cap → moneda sin decimales
        if col_marketcap:
            cell = ws.cell(row=row, column=col_marketcap)
            try:
                valor = float(str(cell.value).replace("$", "").replace(",", ""))
                cell.value = valor
                cell.number_format = '"$"#,##0'
            except:
                pass

    # Autoajuste de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(nombre_archivo)

# ============================================================
# EJECUCIÓN PRINCIPAL
# ============================================================

if __name__ == "__main__":
    print("Cargando CSV desde el Día 02...")
    df = cargar_csv()

    print("Exportando a Excel base...")
    archivo_excel = exportar_a_excel(df)

    print("Aplicando formato profesional...")
    formatear_excel(archivo_excel)

    print("\nArchivo 'informe_cripto.xlsx' generado y formateado correctamente.")
