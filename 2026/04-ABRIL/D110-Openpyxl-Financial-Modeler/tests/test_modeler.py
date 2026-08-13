import pytest
import pandas as pd
import os
import openpyxl
from src.modeler import FinancialModeler

@pytest.fixture
def sample_dataframe():
    return pd.DataFrame({
        "Concepto": ["Ingresos por Ventas", "Costos Operativos"],
        "Mes 1": [4500.0, 1200.0],
        "Mes 2": [6000.0, 1500.0],
        "Mes 3": [5500.0, 1300.0]
    })

def test_generate_model_file(sample_dataframe, tmp_path):
    output_path = tmp_path / "modelo_test.xlsx"
    modeler = FinancialModeler(sample_dataframe)
    result_file = modeler.generate_excel_model(str(output_path))
    
    assert os.path.exists(result_file)
    
    # Validar carga de archivo generado con openpyxl
    wb = openpyxl.load_workbook(result_file)
    assert "Modelo Financiero" in wb.sheetnames
    ws = wb["Modelo Financiero"]
    assert ws["A4"].value == "Ingresos por Ventas"
    # Verificar fórmula de total en columna E
    assert ws["E4"].value.startswith("=SUM")