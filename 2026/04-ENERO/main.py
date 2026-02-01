import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

# ============================================================
# Cargar Excel del Día 03
# ============================================================

def cargar_excel():
    carpeta = os.path.dirname(os.path.abspath(__file__))
    return os.path.normpath(os.path.join(carpeta, "..", "03-ENERO", "informe_cripto.xlsx"))

# ============================================================
# Limpiar datos (formato americano)
# ============================================================

def limpiar_datos(df):
    df["Precio"] = df["Precio"].astype(str).str.replace("$", "").str.replace(",", "")
    df["Precio"] = pd.to_numeric(df["Precio"], errors="coerce")

    df["Market Cap"] = df["Market Cap"].astype(str).str.replace("$", "").str.replace(",", "")
    df["Market Cap"] = pd.to_numeric(df["Market Cap"], errors="coerce")

    df["Cambio 24h"] = df["Cambio 24h"].astype(str).str.replace("%", "")
    df["Cambio 24h"] = pd.to_numeric(df["Cambio 24h"], errors="coerce") / 100

    return df

# ============================================================
# Estilos Premium Ejecutivo
# ============================================================

gris_oscuro = "1C1C1C"
gris_medio = "2A2A2A"
dorado = "D4AF37"
blanco = "FFFFFF"

font_blanca = Font(color=blanco, name="Calibri Light", size=11)
font_dorada = Font(color=dorado, name="Calibri Light", size=12, bold=True)
font_negra = Font(color="000000", name="Calibri Light", size=12, bold=True)

borde_fino = Border(
    left=Side(style="thin", color="666666"),
    right=Side(style="thin", color="666666"),
    top=Side(style="thin", color="666666"),
    bottom=Side(style="thin", color="666666")
)

# ============================================================
# Crear portada ejecutiva
# ============================================================

def crear_portada(wb):
    ws = wb.active
    ws.title = "Portada"

    ws.sheet_view.showGridLines = False

    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=8):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="000000")

    fecha = datetime.now().strftime("%d/%m/%Y")

    ws.merge_cells("A5:H7")
    ws["A5"].value = "MERCADO CRIPTO — ANÁLISIS EJECUTIVO"
    ws["A5"].font = Font(color=blanco, name="Calibri Light", size=28, bold=True)
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A10:H12")
    ws["A10"].value = f"Dashboard Ejecutivo · Actualizado: {fecha}"
    ws["A10"].font = Font(color="CCCCCC", name="Calibri Light", size=16)
    ws["A10"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A20:H22")
    ws["A20"].value = "Oliver Javier Morales Pérez — 365 Python Challenge"
    ws["A20"].font = Font(color=blanco, name="Calibri Light", size=14)
    ws["A20"].alignment = Alignment(horizontal="center", vertical="center")

# ============================================================
# Crear tabla premium
# ============================================================

def crear_tabla(wb, df):
    ws = wb.create_sheet("Tabla")
    ws.sheet_view.showGridLines = False

    columnas = list(df.columns)
    ws.append(columnas)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=dorado)
        cell.font = font_negra
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde_fino

    for fila in df.values.tolist():
        ws.append(fila)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=gris_oscuro)
            cell.font = font_blanca
            cell.border = borde_fino
            cell.alignment = Alignment(horizontal="center")

    # Formatos numéricos corregidos
    col_precio = columnas.index("Precio") + 1
    col_cambio = columnas.index("Cambio 24h") + 1
    col_mc = columnas.index("Market Cap") + 1

    for cell in ws.iter_cols(min_col=col_precio, max_col=col_precio, min_row=2, max_row=ws.max_row):
        for c in cell:
            c.number_format = '#,##0.00'

    for cell in ws.iter_cols(min_col=col_cambio, max_col=col_cambio, min_row=2, max_row=ws.max_row):
        for c in cell:
            c.number_format = '0.00%'

    for cell in ws.iter_cols(min_col=col_mc, max_col=col_mc, min_row=2, max_row=ws.max_row):
        for c in cell:
            c.number_format = '#,##0'

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

# ============================================================
# Crear hoja Market Cap premium
# ============================================================

def crear_marketcap(wb, df):
    ws = wb.create_sheet("MarketCap")
    ws.sheet_view.showGridLines = False

    ws.append(["Criptomoneda", "Market Cap"])
    for _, row in df.iterrows():
        ws.append([row["Nombre"], row["Market Cap"]])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=gris_oscuro)
            cell.font = font_blanca
            cell.border = borde_fino
            cell.alignment = Alignment(horizontal="center")

    # Formato corregido
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '#,##0'

    chart = BarChart()
    chart.title = "Market Cap por Criptomoneda"
    chart.y_axis.title = "USD"
    chart.width = 30
    chart.height = 15

    datos = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    etiquetas = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(datos, titles_from_data=True)
    chart.set_categories(etiquetas)

    ws.add_chart(chart, "A15")

# ============================================================
# Crear hoja de gráficos premium
# ============================================================

def crear_graficos(wb, df):
    ws = wb.create_sheet("Graficos")
    ws.sheet_view.showGridLines = False

    ws.append(["Criptomoneda", "Precio", "Cambio 24h"])
    for _, row in df.iterrows():
        ws.append([row["Nombre"], row["Precio"], row["Cambio 24h"]])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=gris_oscuro)
            cell.font = font_blanca
            cell.border = borde_fino
            cell.alignment = Alignment(horizontal="center")

    # Formatos corregidos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '#,##0.00'

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '0.00%'

    chart_precio = BarChart()
    chart_precio.title = "Precio Actual"
    chart_precio.width = 30
    chart_precio.height = 12

    datos_precio = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    etiquetas = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart_precio.add_data(datos_precio, titles_from_data=True)
    chart_precio.set_categories(etiquetas)

    ws.add_chart(chart_precio, "A15")

    chart_cambio = LineChart()
    chart_cambio.title = "Variación 24h"
    chart_cambio.width = 30
    chart_cambio.height = 12

    datos_cambio = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
    chart_cambio.add_data(datos_cambio, titles_from_data=True)
    chart_cambio.set_categories(etiquetas)

    ws.add_chart(chart_cambio, "A35")

# ============================================================
# Guardar Excel final
# ============================================================

def guardar_excel(wb):
    carpeta = os.path.dirname(os.path.abspath(__file__))
    ruta = os.path.join(carpeta, "informe_cripto.xlsx")
    wb.save(ruta)
    print(f"Archivo generado: {ruta}")

# ============================================================
# Ejecución principal
# ============================================================

if __name__ == "__main__":
    df = pd.read_excel(cargar_excel(), dtype=str)
    df = limpiar_datos(df)

    wb = Workbook()
    crear_portada(wb)
    crear_tabla(wb, df)
    crear_marketcap(wb, df)
    crear_graficos(wb, df)

    guardar_excel(wb)
    print("Día 4 PREMIUM EJECUTIVO completado.")
